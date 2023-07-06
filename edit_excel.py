#Author: MacKayla Smirga
#Date: 6/30/2023
#The purpose of this program is to take an excel sheet full of data and reformat it to look more professional

import tkinter as tk
from tkinter import filedialog, simpledialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from datetime import datetime

class Main:
    def __init__(self):
        self.root = tk.Tk()
        self.root.withdraw()  # Hide the main tkinter window

    def select_file(self):
        # Open file dialog to choose an Excel file
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            client_name = self.get_client_name()
            self.edit_excel(file_path, client_name)
        else:
            print("No file selected.")

    def get_client_name(self):
        # Prompt the user for a client name
        client_name = simpledialog.askstring("Client Name", "Enter the client name:")
        return client_name

    def edit_excel(self, file_path, client_name):
        # Load the workbook
        workbook = load_workbook(filename=file_path)

        # Create a new workbook to save the edited data in. This will be our main workbook now.
        new_workbook = Workbook()

        #Create a counter. This will be used to name the different tables
        x = 1
        
        # Iterate through each sheet in the original workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Create a new sheet in the new workbook with the same name
            new_sheet = new_workbook.create_sheet(sheet_name)
            
            # Iterate over each row in the original sheet and append it to the new sheet.
            #This creates a copy of our original workbook. We can now start editing it.
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

            
            #First, we are going to create a table based off of the data we already have
            
            # Insert 12 lines at the beginning of the new sheet and 1 column
            new_sheet.insert_rows(idx=1, amount=12)
            new_sheet.insert_cols(idx=1, amount=1)

            
            # Determine the range of data in the sheet
            data_range = f"B13:{get_column_letter(new_sheet.max_column)}{new_sheet.max_row}"

            # Create a table from the data range
            table = Table(displayName= f'Table{x}', ref=data_range)
            x += 1

            # Apply a table style
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium21", showFirstColumn=False,
                                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.showAutoFilter = False

            # Add the table to the sheet
            new_sheet.add_table(table)
            
            # Hide the gridlines
            new_sheet.sheet_view.showGridLines = False


            # Adjust Column Widths
            # Find the largest amount of characters in the column, buffer by two, and adjust width accordingly
            #Column B is special, so we will manually updated that one
            for column_cells in new_sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                new_sheet.column_dimensions[column].width = adjusted_width

            new_sheet.column_dimensions['B'].width = 94

            # Set the background color to black, font color to white, and font bold for cell B2
            cell_b2 = new_sheet['B2']
            cell_b2.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell_b2.font = Font(color="FFFFFF", bold=True)

            # Add " / {client_name} /" text to cell B2 along with current month and year
            current_month = datetime.now().strftime("%B")
            current_year = datetime.now().strftime("%Y")
            cell_b2.value = f" / {client_name} / {current_month} {current_year}"

            # Merge and center cells B4 to B9
            merged_range = 'B4:B9'
            new_sheet.merge_cells(merged_range)
            merged_cell = new_sheet[merged_range]
            merged_cell[0][0].alignment = Alignment(horizontal="center", vertical="top")

            # Write "Summary:" in the merged cells. This will be a place to add a summary for each sheet
            merged_cell[0][0].value = "Summary:"

            #Format the summary cells
            #Add a border
            border = Border(left=Side(style='medium', color='636363'),
                            right=Side(style='medium', color='636363'),
                            top=Side(style='medium', color='636363'),
                            bottom=Side(style='medium', color='636363'))
            
            
            for row in range(4, 10):
                cell = new_sheet.cell(row=row, column=2)
                cell.border = border

            # Fill in the current date in cell B10 and edit alignment
            cell_b10 = new_sheet['B10']
            cell_b10.value = f"Date: {datetime.now().date()}"
            cell_b10.alignment = Alignment(horizontal="left", vertical="center")

        #This method creates a blank sheet. Lets get rid of that
        # Remove the first sheet
        new_workbook.remove(new_workbook[new_workbook.sheetnames[0]])
        
        # Save the new workbook with a different filename
        new_filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if new_filename:
            new_workbook.save(filename=new_filename)
            print(f"New Excel file saved as {new_filename}")
        else:
            print("File not saved.")


if __name__ == '__main__':
    main = Main()
    main.select_file()

